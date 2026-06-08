"""
Shared export utilities for weld data (Excel and PDF).
"""
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from django.http import HttpResponse
from django.db.models import Sum, Count


EXPORT_COLUMNS = [
    ("section",          "Section"),
    ("weld_id4",         "Weld ID"),
    ("side",             "Side"),
    ("weld_type",        "Weld Type"),
    ("inspection_stage", "Inspection Stage"),
    ("weld_size",        "Weld Size"),
    ("pass_fail",        "Pass/Fail"),
    ("report",           "Report"),
    ("date",             "Date"),
    ("inspector",        "Inspector"),
    ("repair_welder",    "Repair Welder"),
    ("note",             "Note"),
]


def generate_excel_response(queryset, filename):
    """Return an HttpResponse containing an .xlsx file for the given queryset."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weld Data"

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font_white = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = [col_label for _, col_label in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center

    # Calculate summary aggregates before iterating
    aggregates = queryset.aggregate(
        total_count=Count('id'),
        total_repair_length=Sum('estimated_repair_length'),
        total_weld_length=Sum('total_weld_length'),
    )
    total_count = aggregates['total_count'] or 0
    total_repair_length = aggregates['total_repair_length'] or 0.0
    total_weld_length = aggregates['total_weld_length'] or 0.0

    # Data rows
    for row_idx, weld in enumerate(queryset, start=2):
        for col_idx, (field, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = getattr(weld, field, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=str(value) if not isinstance(value, (int, float)) else value)

    # Auto-width columns
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Summary section
    summary_start_row = ws.max_row + 2
    summary_label_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    summary_label_font = Font(bold=True)

    summary_rows = [
        ("Total Welds", total_count),
        ("Total Repair Length (ft)", round(total_repair_length, 1)),
        ("Total Weld Length (ft)", round(total_weld_length, 1)),
    ]
    for i, (label, value) in enumerate(summary_rows):
        label_cell = ws.cell(row=summary_start_row + i, column=1, value=label)
        label_cell.font = summary_label_font
        label_cell.fill = summary_label_fill
        value_cell = ws.cell(row=summary_start_row + i, column=2, value=value)
        value_cell.font = Font(bold=True)

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def generate_pdf_response(queryset, filename, title, filters_desc=""):
    """Return an HttpResponse containing a landscape PDF for the given queryset."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=4,
    )

    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=8,
        wordWrap='CJK',
    )
    header_cell_style = ParagraphStyle(
        'HeaderCellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=9,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        alignment=1,
    )

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Generated on {date.today().strftime('%B %d, %Y')}", subtitle_style))
    if filters_desc:
        story.append(Paragraph(f"Filters: {filters_desc}", subtitle_style))
    story.append(Spacer(1, 0.4 * cm))

    # Table data
    header_paragraphs = [Paragraph(col_label, header_cell_style) for _, col_label in EXPORT_COLUMNS]
    table_data = [header_paragraphs]

    pass_fill = colors.HexColor("#C6EFCE")
    fail_fill = colors.HexColor("#FFC7CE")

    pass_fail_rows = []  # track row indices (1-based, header is row 0)
    for row_idx, weld in enumerate(queryset, start=1):
        row = []
        for field, _ in EXPORT_COLUMNS:
            value = getattr(weld, field, "") or ""
            row.append(Paragraph(str(value), cell_style))
        table_data.append(row)
        pf = (getattr(weld, "pass_fail", "") or "").lower()
        if pf == "pass":
            pass_fail_rows.append((row_idx, "pass"))
        elif pf == "fail":
            pass_fail_rows.append((row_idx, "fail"))

    col_widths = [2.5*cm, 2*cm, 1.5*cm, 2.5*cm, 3*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 3*cm, 3*cm, 4*cm]

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    # Color-code Pass/Fail column (index 6)
    for row_idx, status in pass_fail_rows:
        fill = pass_fill if status == "pass" else fail_fill
        text_color = colors.HexColor("#276221") if status == "pass" else colors.HexColor("#9C0006")
        style_cmds.append(("BACKGROUND", (6, row_idx), (6, row_idx), fill))
        style_cmds.append(("TEXTCOLOR", (6, row_idx), (6, row_idx), text_color))
        style_cmds.append(("FONTNAME", (6, row_idx), (6, row_idx), "Helvetica-Bold"))

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    # Summary section
    aggregates = queryset.aggregate(
        total_count=Count('id'),
        total_repair_length=Sum('estimated_repair_length'),
        total_weld_length=Sum('total_weld_length'),
    )
    total_count = aggregates['total_count'] or 0
    total_repair_length = aggregates['total_repair_length'] or 0.0
    total_weld_length = aggregates['total_weld_length'] or 0.0

    story.append(Spacer(1, 0.5 * cm))
    summary_style = ParagraphStyle(
        "SummaryLabel",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=2,
    )
    summary_data = [
        [Paragraph("<b>Summary</b>", summary_style), ""],
        [Paragraph("Total Welds", summary_style), Paragraph(str(total_count), summary_style)],
        [Paragraph("Total Repair Length (ft)", summary_style), Paragraph(f"{total_repair_length:.1f}", summary_style)],
        [Paragraph("Total Weld Length (ft)", summary_style), Paragraph(f"{total_weld_length:.1f}", summary_style)],
    ]
    summary_tbl = Table(summary_data, colWidths=[5 * cm, 3 * cm])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story.append(summary_tbl)

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
