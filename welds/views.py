import json
import io
import re
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q, Max
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from welds.models import Weld, WeldPhoto, WeldIdKey
from welds.export_utils import generate_excel_response, generate_pdf_response

# Fields that may be edited via the inline-edit API
_WELD_EDITABLE_FIELDS = {
    'report', 'side', 'section', 'weld_id', 'weld_id2', 'weld_id3', 'weld_id4',
    'estimated_repair_length', 'total_weld_length',
    'table_6_1_criteria_1', 'table_6_1_criteria_2', 'table_6_1_criteria_3',
    'weld_type', 'weld_size', 'wps_number',
    'inspection_utsw', 'inspection_mt', 'inspector', 'date',
    'pass_fail', 'corrective_action_taken', 'repair_welder',
    'repair_inspection_date', 'weld_process', 'note',
    'validation_note', 'validation_cleared',
}


def _apply_weld_filters(request):
    """Apply weld_list GET filters and return a queryset."""
    side = request.GET.get('side')
    section = request.GET.get('section')
    weld_id = request.GET.get('weld_id')
    weld_type = request.GET.get('weld_type')
    pass_fail = request.GET.get('pass_fail')
    report = request.GET.get('report')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    inspector = request.GET.get('inspector')
    inspection_stage = request.GET.get('inspection_stage')

    queryset = Weld.objects.all()
    if side:
        queryset = queryset.filter(side=side)
    if section:
        queryset = queryset.filter(section=section)
    if weld_id:
        queryset = queryset.filter(weld_id__icontains=weld_id)
    if weld_type:
        queryset = queryset.filter(weld_type=weld_type)
    if pass_fail:
        queryset = queryset.filter(pass_fail=pass_fail)
    if report:
        queryset = queryset.filter(report=report)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    if inspector:
        queryset = queryset.filter(inspector=inspector)
    if inspection_stage:
        queryset = queryset.filter(inspection_stage=inspection_stage)
    if search:
        queryset = queryset.filter(
            Q(section__icontains=search) |
            Q(weld_id__icontains=search) |
            Q(weld_id4__icontains=search) |
            Q(inspector__icontains=search) |
            Q(note__icontains=search)
        )
    return queryset


def _build_filters_desc(request):
    """Return a human-readable string describing active filters."""
    parts = []
    for key, label in [('side', 'Side'), ('section', 'Section'), ('weld_type', 'Type'),
                       ('pass_fail', 'Status'), ('report', 'Report'), ('inspector', 'Inspector'),
                       ('inspection_stage', 'Stage'),
                       ('date_from', 'From'), ('date_to', 'To'), ('search', 'Search')]:
        val = request.GET.get(key)
        if val:
            parts.append(f"{label}: {val}")
    return ", ".join(parts) if parts else "None"


@login_required
def weld_list(request):
    queryset = _apply_weld_filters(request)

    # Calculate aggregates, case-insensitive for pass/fail
    aggregates = queryset.aggregate(
        total_count=Count('id'),
        total_repair_length=Sum('estimated_repair_length'),
        total_weld_length=Sum('total_weld_length'),
        pass_count=Count('id', filter=Q(pass_fail__iexact='pass')),
        fail_count=Count('id', filter=Q(pass_fail__iexact='fail'))
    )

    # Get distinct values for filter dropdowns
    sides = Weld.objects.values_list('side', flat=True).distinct()
    weld_types = Weld.objects.values_list('weld_type', flat=True).distinct()
    reports = Weld.objects.values_list('report', flat=True).distinct()
    pass_fail_choices = Weld.objects.values_list('pass_fail', flat=True).distinct()
    sections = Weld.objects.values_list('section', flat=True).order_by('section').distinct()
    inspectors = Weld.objects.values_list('inspector', flat=True).order_by('inspector').distinct().exclude(inspector='')
    stages = Weld.objects.values_list('inspection_stage', flat=True).order_by('inspection_stage').distinct().exclude(inspection_stage='')

    # Paginate results
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'welds': page_obj.object_list,
        'aggregates': aggregates,
        'sides': sides,
        'weld_types': weld_types,
        'reports': reports,
        'pass_fail_choices': pass_fail_choices,
        'sections': sections,
        'inspectors': inspectors,
        'stages': stages,
        'selected_stage': request.GET.get('inspection_stage', ''),
    }

    return render(request, 'welds/weld_list.html', context)


@login_required
def dashboard(request):
    from reports.models import NDEReport

    total_welds = Weld.objects.count()
    pass_count = Weld.objects.filter(pass_fail__iexact='pass').count()
    fail_count = Weld.objects.filter(pass_fail__iexact='fail').count()
    incomplete_count = Weld.objects.filter(validation_cleared=False).filter(_INCOMPLETE_Q).count()
    not_inspected = Weld.objects.filter(pass_fail='').count()

    inspected = pass_count + fail_count
    pass_rate = round((pass_count / inspected) * 100, 1) if inspected > 0 else 0

    section_data = (
        Weld.objects.values('section')
        .annotate(
            total=Count('id'),
            passed=Count('id', filter=Q(pass_fail__iexact='pass')),
            failed=Count('id', filter=Q(pass_fail__iexact='fail')),
        )
        .order_by('section')
    )

    total_photos = WeldPhoto.objects.count()
    total_reports = NDEReport.objects.count()
    recent_welds = Weld.objects.order_by('-updated_at')[:5]

    donut_data = {
        'pass': pass_count,
        'fail': fail_count,
        'not_inspected': not_inspected,
    }
    section_labels = [s['section'] for s in section_data]
    section_pass = [s['passed'] for s in section_data]
    section_fail = [s['failed'] for s in section_data]

    return render(request, 'welds/dashboard.html', {
        'total_welds': total_welds,
        'pass_count': pass_count,
        'fail_count': fail_count,
        'incomplete_count': incomplete_count,
        'not_inspected': not_inspected,
        'pass_rate': pass_rate,
        'total_photos': total_photos,
        'total_reports': total_reports,
        'recent_welds': recent_welds,
        'donut_data_json': json.dumps(donut_data),
        'section_labels_json': json.dumps(section_labels),
        'section_pass_json': json.dumps(section_pass),
        'section_fail_json': json.dumps(section_fail),
    })


@login_required
def weld_detail(request, pk):
    # Get single weld
    weld = get_object_or_404(Weld, pk=pk)

    # Get previous and next weld by pk
    previous_weld = Weld.objects.filter(pk__lt=pk).order_by('-pk').first()
    next_weld = Weld.objects.filter(pk__gt=pk).order_by('pk').first()

    # Decode Weld ID4 segments against WeldIdKey lookup table
    weld_id4_decoded = _decode_weld_id4(weld.weld_id4)

    context = {
        'weld': weld,
        'previous_weld': previous_weld,
        'next_weld': next_weld,
        'weld_id4_decoded': weld_id4_decoded,
    }

    return render(request, 'welds/weld_detail.html', context)


@login_required
def export_welds_excel(request):
    queryset = _apply_weld_filters(request)
    filename = f"weld_data_export_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return generate_excel_response(queryset, filename)


@login_required
def export_welds_pdf(request):
    queryset = _apply_weld_filters(request)
    filename = f"weld_report_{date.today().strftime('%Y-%m-%d')}.pdf"
    filters_desc = _build_filters_desc(request)
    return generate_pdf_response(
        queryset,
        filename,
        title="TVA Barkley Dam \u2014 Weld Inspection Report",
        filters_desc=filters_desc,
    )


@login_required
@require_POST
def weld_update(request, pk):
    """AJAX endpoint: update editable fields on a Weld record."""
    weld = get_object_or_404(Weld, pk=pk)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    for field, value in data.items():
        if field not in _WELD_EDITABLE_FIELDS:
            return JsonResponse(
                {'success': False, 'error': f'Field "{field}" is not editable'},
                status=400,
            )
        field_meta = Weld._meta.get_field(field)
        # Convert empty string to None for nullable fields
        if value == '' and getattr(field_meta, 'null', False):
            value = None
        setattr(weld, field, value)

    try:
        weld.full_clean()
        weld.save()
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Weld ID Key helper
# ---------------------------------------------------------------------------

def _decode_weld_id4(weld_id4):
    """
    Split the weld_id4 string on hyphens, underscores, or spaces and
    look up each segment in the WeldIdKey table.
    Returns a list of (segment, meaning_or_None) tuples.
    """
    if not weld_id4:
        return []
    segments = [s for s in re.split(r'[-_ ]+', weld_id4) if s]
    if not segments:
        return []
    key_map = {k.code: k.meaning for k in WeldIdKey.objects.filter(code__in=segments)}
    return [(seg, key_map.get(seg)) for seg in segments]


# ---------------------------------------------------------------------------
# Weld ID Key views
# ---------------------------------------------------------------------------

@login_required
def weld_id_key(request):
    """Display and manage Weld ID Key entries."""
    entries = WeldIdKey.objects.all()
    return render(request, 'welds/weld_id_key.html', {'entries': entries})


@login_required
@require_POST
def weld_id_key_create(request):
    """AJAX: create a new WeldIdKey entry."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    code = data.get('code', '').strip()
    meaning = data.get('meaning', '').strip()
    if not code or not meaning:
        return JsonResponse({'success': False, 'error': 'Both code and meaning are required.'}, status=400)
    if len(code) > 10:
        return JsonResponse({'success': False, 'error': 'Code must be 10 characters or fewer.'}, status=400)
    if len(meaning) > 100:
        return JsonResponse({'success': False, 'error': 'Meaning must be 100 characters or fewer.'}, status=400)

    if WeldIdKey.objects.filter(code=code).exists():
        return JsonResponse({'success': False, 'error': f'Code "{code}" already exists.'}, status=400)

    entry = WeldIdKey.objects.create(code=code, meaning=meaning)
    return JsonResponse({'success': True, 'id': entry.pk, 'code': entry.code, 'meaning': entry.meaning})


@login_required
@require_POST
def weld_id_key_update(request, pk):
    """AJAX: update an existing WeldIdKey entry."""
    entry = get_object_or_404(WeldIdKey, pk=pk)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    code = data.get('code', '').strip()
    meaning = data.get('meaning', '').strip()
    if not code or not meaning:
        return JsonResponse({'success': False, 'error': 'Both code and meaning are required.'}, status=400)
    if len(code) > 10:
        return JsonResponse({'success': False, 'error': 'Code must be 10 characters or fewer.'}, status=400)
    if len(meaning) > 100:
        return JsonResponse({'success': False, 'error': 'Meaning must be 100 characters or fewer.'}, status=400)

    if WeldIdKey.objects.filter(code=code).exclude(pk=pk).exists():
        return JsonResponse({'success': False, 'error': f'Code "{code}" already exists.'}, status=400)

    entry.code = code
    entry.meaning = meaning
    entry.save()
    return JsonResponse({'success': True, 'id': entry.pk, 'code': entry.code, 'meaning': entry.meaning})


@login_required
@require_POST
def weld_id_key_delete(request, pk):
    """AJAX: delete a WeldIdKey entry."""
    entry = get_object_or_404(WeldIdKey, pk=pk)
    entry.delete()
    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Incomplete records helpers
# ---------------------------------------------------------------------------

_INCOMPLETE_Q = (
    Q(inspector='') | Q(date__isnull=True) | Q(pass_fail='') |
    Q(weld_type='') | Q(total_weld_length__isnull=True)
)


def _apply_incomplete_filters(section='', report='', missing_field='', search=''):
    """Return a queryset of not-cleared welds with at least one missing critical field,
    with optional extra filters applied."""
    queryset = Weld.objects.filter(validation_cleared=False).filter(_INCOMPLETE_Q)
    if section:
        queryset = queryset.filter(section=section)
    if report:
        queryset = queryset.filter(report=report)
    if missing_field == 'inspector':
        queryset = queryset.filter(inspector='')
    elif missing_field == 'date':
        queryset = queryset.filter(date__isnull=True)
    elif missing_field == 'pass_fail':
        queryset = queryset.filter(pass_fail='')
    elif missing_field == 'weld_type':
        queryset = queryset.filter(weld_type='')
    elif missing_field == 'total_weld_length':
        queryset = queryset.filter(total_weld_length__isnull=True)
    if search:
        queryset = queryset.filter(
            Q(section__icontains=search) |
            Q(weld_id4__icontains=search) |
            Q(inspector__icontains=search)
        )
    return queryset.order_by('section', 'weld_id4')


# ---------------------------------------------------------------------------
# Incomplete records views
# ---------------------------------------------------------------------------

@login_required
def incomplete_records(request):
    """Page showing welds with missing critical fields that haven't been cleared."""
    section = request.GET.get('section', '')
    report = request.GET.get('report', '')
    missing_field = request.GET.get('missing_field', '')
    search = request.GET.get('search', '')

    queryset = _apply_incomplete_filters(section, report, missing_field, search)
    total_count = queryset.count()

    # Dropdown data from all incomplete welds (unfiltered by section/report/field)
    all_incomplete = Weld.objects.filter(validation_cleared=False).filter(_INCOMPLETE_Q)
    sections = all_incomplete.values_list('section', flat=True).order_by('section').distinct()
    reports = all_incomplete.values_list('report', flat=True).order_by('report').distinct()

    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'welds/incomplete_records.html', {
        'page_obj': page_obj,
        'welds': page_obj.object_list,
        'total_count': total_count,
        'sections': sections,
        'reports': reports,
        'selected_section': section,
        'selected_report': report,
        'selected_missing_field': missing_field,
        'search': search,
    })


@login_required
def export_incomplete_excel(request):
    """Export filtered incomplete records as an .xlsx file with a Missing Fields column."""
    section = request.GET.get('section', '')
    report = request.GET.get('report', '')
    missing_field = request.GET.get('missing_field', '')
    search = request.GET.get('search', '')

    queryset = _apply_incomplete_filters(section, report, missing_field, search)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incomplete Records"

    headers = ["Section", "Weld ID4", "Report", "Inspector", "Date",
               "Pass/Fail", "Weld Type", "Total Weld Length", "Missing Fields"]
    header_fill = PatternFill(start_color="C85E00", end_color="C85E00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for weld in queryset:
        missing = weld.get_missing_fields()
        ws.append([
            weld.section,
            weld.weld_id4,
            weld.report,
            weld.inspector,
            weld.date.strftime('%Y-%m-%d') if weld.date else '',
            weld.pass_fail,
            weld.weld_type,
            weld.total_weld_length,
            ', '.join(missing),
        ])

    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 60)

    filename = f"incomplete_records_{date.today().strftime('%Y-%m-%d')}.xlsx"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def weld_clear_validation(request, pk):
    """AJAX: clear the validation warning for a weld by providing a note."""
    weld = get_object_or_404(Weld, pk=pk)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    note = data.get('note', '').strip()
    if not note:
        return JsonResponse({'success': False, 'error': 'A note is required to clear the warning.'}, status=400)

    weld.validation_note = note
    weld.validation_cleared = True
    weld.save(update_fields=['validation_note', 'validation_cleared', 'updated_at'])
    return JsonResponse({'success': True})


@login_required
@require_POST
def weld_unclear_validation(request, pk):
    """AJAX: remove the validation-cleared status from a weld."""
    weld = get_object_or_404(Weld, pk=pk)
    weld.validation_note = ''
    weld.validation_cleared = False
    weld.save(update_fields=['validation_note', 'validation_cleared', 'updated_at'])
    return JsonResponse({'success': True})


@login_required
@require_POST
def weld_bulk_clear_validation(request):
    """AJAX: clear validation warnings for a list of weld PKs."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    note = data.get('note', '').strip()
    pks = data.get('pks', [])

    if not note:
        return JsonResponse({'success': False, 'error': 'A note is required to clear the warnings.'}, status=400)
    if not pks:
        return JsonResponse({'success': False, 'error': 'No welds selected.'}, status=400)

    count = Weld.objects.filter(pk__in=pks).update(
        validation_note=note,
        validation_cleared=True,
        updated_at=timezone.now(),
    )
    return JsonResponse({'success': True, 'count': count})


@login_required
@require_POST
def weld_bulk_clear_all_filtered(request):
    """AJAX: clear validation warnings for ALL welds matching the current filters."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    note = data.get('note', '').strip()
    if not note:
        return JsonResponse({'success': False, 'error': 'A note is required.'}, status=400)

    section = data.get('section', '')
    report = data.get('report', '')
    missing_field = data.get('missing_field', '')
    search = data.get('search', '')

    queryset = _apply_incomplete_filters(section, report, missing_field, search)
    count = queryset.update(validation_note=note, validation_cleared=True, updated_at=timezone.now())
    return JsonResponse({'success': True, 'count': count})


# ---------------------------------------------------------------------------
# Section Map view
# ---------------------------------------------------------------------------

def _build_section_stats(pass_count, fail_repaired, fail_unrepaired, incomplete_count):
    """Return a human-readable stats string for a section tile, e.g. '45 Pass · 3 Repaired · 1 Fail · 2 Incomplete'."""
    parts = []
    if pass_count:
        parts.append(f'{pass_count} Pass')
    if fail_repaired:
        parts.append(f'{fail_repaired} Repaired')
    if fail_unrepaired:
        parts.append(f'{fail_unrepaired} Fail')
    if incomplete_count:
        parts.append(f'{incomplete_count} Incomplete')
    return ' · '.join(parts)


@login_required
def section_map(request):
    """Section map showing color-coded tiles for each section based on inspection status."""
    side_filter = request.GET.get('side', '')

    qs = Weld.objects.all()
    if side_filter:
        qs = qs.filter(side=side_filter)

    # Aggregate stats per section using a single DB query
    sections_qs = qs.values('section').annotate(
        side=Max('side'),
        total=Count('id'),
        pass_count=Count('id', filter=Q(pass_fail='Pass')),
        fail_unrepaired=Count('id', filter=Q(pass_fail='Fail') & Q(repair_inspection_date__isnull=True)),
        fail_repaired=Count('id', filter=Q(pass_fail='Fail') & Q(repair_inspection_date__isnull=False)),
        incomplete_count=Count('id', filter=_INCOMPLETE_Q & Q(validation_cleared=False)),
        not_inspected_count=Count('id', filter=Q(pass_fail='')),
    ).order_by('section')

    # Determine worst-case status for each tile
    section_tiles = []
    for s in sections_qs:
        if s['fail_unrepaired'] > 0:
            status = 'red'
            status_label = 'Has Failures'
        elif s['incomplete_count'] > 0:
            status = 'amber'
            status_label = 'Incomplete'
        elif s['fail_repaired'] > 0:
            status = 'blue'
            status_label = 'All Repaired'
        elif s['not_inspected_count'] == 0 and s['pass_count'] > 0:
            status = 'green'
            status_label = 'All Passed'
        else:
            status = 'gray'
            status_label = 'Not Inspected'

        section_tiles.append({
            'section': s['section'],
            'side': s['side'] or '',
            'total': s['total'],
            'pass_count': s['pass_count'],
            'fail_unrepaired': s['fail_unrepaired'],
            'fail_repaired': s['fail_repaired'],
            'incomplete_count': s['incomplete_count'],
            'not_inspected_count': s['not_inspected_count'],
            'status': status,
            'status_label': status_label,
            'stats_summary': _build_section_stats(
                s['pass_count'], s['fail_repaired'],
                s['fail_unrepaired'], s['incomplete_count'],
            ),
        })

    # Summary stats
    total_sections = len(section_tiles)
    all_passed = sum(1 for s in section_tiles if s['status'] == 'green')
    has_failures = sum(1 for s in section_tiles if s['status'] == 'red')
    all_repaired = sum(1 for s in section_tiles if s['status'] == 'blue')
    has_incomplete = sum(1 for s in section_tiles if s['status'] == 'amber')
    not_inspected_summary = sum(1 for s in section_tiles if s['status'] == 'gray')

    # Distinct sides for the filter dropdown
    all_sides = (
        Weld.objects.exclude(side='')
        .values_list('side', flat=True)
        .order_by('side')
        .distinct()
    )

    # Group tiles by side for optional side-grouped display
    sides_groups = {}
    for tile in section_tiles:
        side = tile['side'] or 'Unknown'
        sides_groups.setdefault(side, []).append(tile)

    return render(request, 'welds/section_map.html', {
        'section_tiles': section_tiles,
        'sides_groups': sides_groups,
        'total_sections': total_sections,
        'all_passed': all_passed,
        'has_failures': has_failures,
        'all_repaired': all_repaired,
        'has_incomplete': has_incomplete,
        'not_inspected_summary': not_inspected_summary,
        'all_sides': all_sides,
        'selected_side': side_filter,
    })
