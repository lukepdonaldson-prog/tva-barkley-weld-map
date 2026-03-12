import io
import json
import zipfile
from datetime import date

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from welds.models import Weld, WeldPhoto
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from welds.export_utils import generate_excel_response, generate_pdf_response

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Fields that may be edited via the inline-edit API
_PHOTO_EDITABLE_FIELDS = {'description', 'section'}


def _apply_qa_filters(request):
    """Apply qa_dashboard GET filters and return a queryset."""
    welds = Weld.objects.all()
    section = request.GET.get('section', '')
    report = request.GET.get('report', '')
    pass_fail = request.GET.get('pass_fail', '')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    inspector = request.GET.get('inspector', '')

    if section:
        welds = welds.filter(section=section)
    if report:
        welds = welds.filter(report=report)
    if pass_fail:
        welds = welds.filter(pass_fail=pass_fail)
    if date_from:
        welds = welds.filter(date__gte=date_from)
    if date_to:
        welds = welds.filter(date__lte=date_to)
    if inspector:
        welds = welds.filter(inspector=inspector)
    if search:
        welds = welds.filter(
            Q(weld_id4__icontains=search) |
            Q(inspector__icontains=search) |
            Q(note__icontains=search)
        )
    return welds.order_by('section', 'weld_id4')


def _build_qa_filters_desc(request):
    """Return a human-readable string describing active QA dashboard filters."""
    parts = []
    for key, label in [('section', 'Section'), ('report', 'Report'),
                       ('pass_fail', 'Status'), ('inspector', 'Inspector'),
                       ('date_from', 'From'), ('date_to', 'To'), ('search', 'Search')]:
        val = request.GET.get(key)
        if val:
            parts.append(f"{label}: {val}")
    return ", ".join(parts) if parts else "None"


@login_required
def qa_dashboard(request):
    welds = _apply_qa_filters(request)

    section = request.GET.get('section', '')
    report = request.GET.get('report', '')
    pass_fail = request.GET.get('pass_fail', '')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    inspector = request.GET.get('inspector', '')

    count_all = Weld.objects.count()
    count_filtered = welds.count()

    pass_count = welds.filter(pass_fail__iexact='pass').count()
    fail_count = welds.filter(pass_fail__iexact='fail').count()

    # Count of incomplete welds (missing critical fields, not cleared) matching current filters
    from django.db.models import Q as _Q
    _incomplete_q = (
        _Q(inspector='') | _Q(date__isnull=True) | _Q(pass_fail='') |
        _Q(weld_type='') | _Q(total_weld_length__isnull=True)
    )
    incomplete_count = welds.filter(validation_cleared=False).filter(_incomplete_q).count()

    paginator = Paginator(welds, 50)
    page_number = request.GET.get('page')
    welds_page = paginator.get_page(page_number)

    sections = Weld.objects.values_list('section', flat=True).order_by('section').distinct()
    reports = Weld.objects.values_list('report', flat=True).order_by('report').distinct()
    statuses = Weld.objects.values_list('pass_fail', flat=True).order_by('pass_fail').distinct().exclude(pass_fail='')
    inspectors = Weld.objects.values_list('inspector', flat=True).order_by('inspector').distinct().exclude(inspector='')

    return render(request, 'gallery/qa_dashboard.html', {
        'welds': welds_page,
        'count_all': count_all,
        'count_filtered': count_filtered,
        'pass_count': pass_count,
        'fail_count': fail_count,
        'incomplete_count': incomplete_count,
        'sections': sections,
        'reports': reports,
        'statuses': statuses,
        'inspectors': inspectors,
        'selected_section': section,
        'selected_report': report,
        'selected_status': pass_fail,
        'search': search,
        'selected_inspector': inspector,
        'date_from': date_from,
        'date_to': date_to,
    })

def _apply_photo_filters(request):
    """Apply photo gallery GET filters and return a queryset."""
    photos = WeldPhoto.objects.all()
    section = request.GET.get('section', '')
    report = request.GET.get('report_number', '')
    subfolder = request.GET.get('subfolder', '')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if section:
        photos = photos.filter(section__icontains=section)
    if report:
        photos = photos.filter(report_number__icontains=report)
    if subfolder:
        photos = photos.filter(subfolder__icontains=subfolder)
    if date_from:
        photos = photos.filter(uploaded_at__date__gte=date_from)
    if date_to:
        photos = photos.filter(uploaded_at__date__lte=date_to)
    if search:
        photos = photos.filter(
            Q(description__icontains=search) |
            Q(original_filename__icontains=search)
        )
    return photos.order_by('-uploaded_at')


@login_required
def photo_gallery(request):
    photos = _apply_photo_filters(request)

    section = request.GET.get('section', '')
    report = request.GET.get('report_number', '')
    subfolder = request.GET.get('subfolder', '')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    paginator = Paginator(photos, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    sections = WeldPhoto.objects.values_list('section', flat=True).order_by('section').distinct()
    reports = WeldPhoto.objects.values_list('report_number', flat=True).order_by('report_number').distinct()
    subfolders = WeldPhoto.objects.values_list('subfolder', flat=True).order_by('subfolder').distinct()

    return render(request, 'gallery/gallery.html', {
        'page_obj': page_obj,
        'count_all': WeldPhoto.objects.count(),
        'count_filtered': photos.count(),
        'sections': sections,
        'reports': reports,
        'subfolders': subfolders,
        'selected_section': section,
        'selected_report': report,
        'selected_subfolder': subfolder,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
def export_qa_excel(request):
    queryset = _apply_qa_filters(request)
    filename = f"qa_dashboard_export_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return generate_excel_response(queryset, filename)


@login_required
def export_qa_pdf(request):
    queryset = _apply_qa_filters(request)
    filename = f"qa_dashboard_report_{date.today().strftime('%Y-%m-%d')}.pdf"
    filters_desc = _build_qa_filters_desc(request)
    return generate_pdf_response(
        queryset,
        filename,
        title="TVA Barkley Dam \u2014 Weld Inspection Report",
        filters_desc=filters_desc,
    )


@login_required
def export_photos_excel(request):
    """Export filtered photo metadata as an .xlsx file."""
    photos = _apply_photo_filters(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Photo Library"

    headers = ["Section", "Report #", "Subfolder", "Original Filename", "Description", "Uploaded Date"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for photo in photos:
        ws.append([
            photo.section,
            photo.report_number,
            photo.subfolder,
            photo.original_filename,
            photo.description,
            photo.uploaded_at.strftime('%Y-%m-%d %H:%M') if photo.uploaded_at else '',
        ])

    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 60)

    filename = f"photo_library_export_{date.today().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_photos_zip(request):
    """Download all filtered photos as a ZIP file."""
    photos = _apply_photo_filters(request)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for photo in photos:
            try:
                file_path = photo.photo.path
                archive_name = f"{photo.section}/{photo.subfolder}/{photo.original_filename}"
                zf.write(file_path, archive_name)
            except (FileNotFoundError, OSError):
                continue

    buffer.seek(0)
    filename = f"weld_photos_{date.today().strftime('%Y-%m-%d')}.zip"
    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def photo_update(request, pk):
    """AJAX endpoint: update editable fields on a WeldPhoto record."""
    photo = get_object_or_404(WeldPhoto, pk=pk)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    for field, value in data.items():
        if field not in _PHOTO_EDITABLE_FIELDS:
            return JsonResponse(
                {'success': False, 'error': f'Field "{field}" is not editable'},
                status=400,
            )
        setattr(photo, field, value)

    try:
        photo.save()
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    return JsonResponse({'success': True})
